"""
匯出功能路由
"""
from flask import Blueprint, request, jsonify, send_file
from services.excel_service import ExcelService
from services.word_service import WordService
from services.google_maps_template_service import generate_google_maps_style_html
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime
from utils.path_manager import get_output_dir
from pathlib import Path
import os

bp = Blueprint('export', __name__)
excel_service = ExcelService()
word_service = WordService()


@bp.route('/excel', methods=['POST'])
def export_excel():
    """
    匯出更新後的 Excel 檔案
    
    請求:
        {
            "file_path": "原始檔案路徑",
            "records": 包含計算結果的紀錄列表
        }
    
    回應:
        Excel 檔案下載
    """
    try:
        data = request.get_json()
        file_path = data.get('file_path')
        records = data.get('records', [])
        
        if not file_path or not os.path.exists(file_path):
            return jsonify({
                'status': 'error',
                'message': '原始檔案不存在'
            }), 400
        
        if not records:
            return jsonify({
                'status': 'error',
                'message': '沒有資料可匯出'
            }), 400
        
        # 更新 Excel 檔案
        output_path = excel_service.add_calculation_results(file_path, records)
        
        # 回傳檔案
        return send_file(
            output_path,
            as_attachment=True,
            download_name=os.path.basename(output_path),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"匯出 Excel 錯誤: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'匯出失敗: {str(e)}'
        }), 500


@bp.route('/word', methods=['POST'])
def export_word():
    """
    匯出 Word 報表（依計畫別）
    
    請求:
        {
            "project_name": "計畫別名稱",
            "records": 該計畫別的紀錄列表（含計算結果）,
            "fixed_origin": "固定起點地址（可選）"
        }
    
    回應:
        Word 檔案下載
    """
    try:
        data = request.get_json()
        project_name = data.get('project_name', '未分類')
        records = data.get('records', [])
        fixed_origin = data.get('fixed_origin', '')
        
        if not records:
            return jsonify({
                'status': 'error',
                'message': '沒有資料可匯出'
            }), 400
        
        # 產生 Word 報表
        word_path = word_service.generate_report(project_name, records, fixed_origin)
        
        # 回傳檔案
        return send_file(
            word_path,
            as_attachment=True,
            download_name=os.path.basename(word_path),
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        
    except Exception as e:
        logger.error(f"匯出 Word 錯誤: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'匯出失敗: {str(e)}'
        }), 500


@bp.route('/word/batch', methods=['POST'])
def export_word_batch():
    """
    批次匯出多個計畫別的 Word 報表
    
    請求:
        {
            "projects": {
                "計畫別名稱1": [records...],
                "計畫別名稱2": [records...],
                ...
            },
            "fixed_origin": "固定起點地址（可選）"
        }
    
    回應:
        ZIP 壓縮檔（包含所有 Word 檔案）
    """
    try:
        data = request.get_json()
        projects = data.get('projects', {})
        fixed_origin = data.get('fixed_origin', '')
        
        if not projects:
            return jsonify({
                'status': 'error',
                'message': '沒有資料可匯出'
            }), 400
        
        # 產生所有 Word 報表
        word_files = []
        for project_name, records in projects.items():
            try:
                word_path = word_service.generate_report(project_name, records, fixed_origin)
                word_files.append({
                    'path': word_path,
                    'name': os.path.basename(word_path)
                })
            except Exception as e:
                logger.error(f"產生 {project_name} 報表錯誤: {str(e)}")
                continue
        
        if not word_files:
            return jsonify({
                'status': 'error',
                'message': '無法產生任何報表'
            }), 500
        
        # 建立 ZIP 壓縮檔（使用相對路徑）
        import zipfile
        from datetime import datetime
        
        output_dir = get_output_dir()
        
        zip_filename = f"里程報表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        zip_path = output_dir / zip_filename
        
        with zipfile.ZipFile(str(zip_path), 'w', zipfile.ZIP_DEFLATED) as zipf:
            for word_file in word_files:
                zipf.write(word_file['path'], word_file['name'])
        
        logger.info(f"成功產生 ZIP 壓縮檔: {zip_path}, 包含 {len(word_files)} 個 Word 檔案")
        
        # 回傳 ZIP 檔案
        return send_file(
            zip_path,
            as_attachment=True,
            download_name=zip_filename,
            mimetype='application/zip'
        )
        
    except Exception as e:
        logger.error(f"批次匯出 Word 錯誤: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'匯出失敗: {str(e)}'
        }), 500


@bp.route('/template', methods=['GET'])
def export_template():
    """
    下載 Excel 範本檔案
    
    回應:
        Excel 範本檔案下載
    """
    try:
        # 建立新的 Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "範本"
        
        # 設定欄位名稱（對齊系統解析所需欄位）
        headers = [
            '部門',
            '姓名',
            '計畫別',
            '起點名稱',
            '出差日期時間（開始）',
            '出差日期時間（結束）',
            '目的地名稱'
        ]
        
        # 設定標題列樣式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 寫入標題列
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 設定標題列高度
        ws.row_dimensions[1].height = 25
        
        # 寫入示例資料（第二列）
        example_data = [
            '安環處',
            '張三',
            'IDA智慧工安',
            '安環高雄處',
            '2024-10-22T09:00:00',  # ISO 格式
            '2024-10-22T17:00:00',  # ISO 格式
            '經濟部產業園區管理局'
        ]
        
        for col_idx, value in enumerate(example_data, start=1):
            cell = ws.cell(row=2, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 設定欄寬
        column_widths = {
            'A': 15,  # 部門
            'B': 12,  # 姓名
            'C': 20,  # 計畫別
            'D': 25,  # 起點名稱
            'E': 25,  # 出差日期時間（開始）
            'F': 25,  # 出差日期時間（結束）
            'G': 30   # 目的地名稱
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 將 Workbook 寫入 BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info("成功產生 Excel 範本檔案")
        
        # 回傳檔案
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='里程報表範本.xlsx'
        )
        
    except Exception as e:
        logger.error(f"產生 Excel 範本錯誤: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'產生範本失敗: {str(e)}'
        }), 500


@bp.route('/html', methods=['POST'])
def export_html():
    """
    匯出 Google Maps 風格的 HTML 檔案（單筆記錄）
    
    請求:
        {
            "record": 單筆記錄（含計算結果）,
            "fixed_origin": "固定起點地址（可選）"
        }
    
    回應:
        HTML 檔案下載
    """
    try:
        data = request.get_json()
        record = data.get('record')
        fixed_origin = data.get('fixed_origin', '')
        
        if not record:
            return jsonify({
                'status': 'error',
                'message': '沒有資料可匯出'
            }), 400
        
        # 產生 Google Maps 風格 HTML
        html_path = generate_google_maps_style_html(
            record=record,
            output_path=None,  # 自動生成檔名
            fixed_origin=fixed_origin
        )
        
        # 回傳檔案
        return send_file(
            html_path,
            as_attachment=True,
            download_name=os.path.basename(html_path),
            mimetype='text/html'
        )
        
    except Exception as e:
        logger.error(f"匯出 HTML 錯誤: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'匯出失敗: {str(e)}'
        }), 500

