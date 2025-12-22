"""
報表產生工具
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from datetime import datetime
import os
import requests
from loguru import logger

class ExcelReportGenerator:
    """Excel 報表產生器"""
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "里程報表"
    
    def generate_mileage_report(self, records, report_type='detail', include_map=False, include_route=False, include_distance=True):
        """產生里程報表"""
        try:
            # 設定標題
            self.ws['A1'] = '地點里程報表'
            self.ws['A1'].font = Font(size=16, bold=True)
            self.ws.merge_cells('A1:G1')
            
            # 設定表頭
            headers = ['日期', '起點', '終點', '單程(km)', '往返(km)', '預估時間', '路線說明']
            for col, header in enumerate(headers, start=1):
                cell = self.ws.cell(row=2, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 填入資料
            for row, record in enumerate(records, start=3):
                self.ws.cell(row=row, column=1).value = record.get('travel_date')
                self.ws.cell(row=row, column=2).value = record.get('start_location')
                self.ws.cell(row=row, column=3).value = record.get('end_location')
                self.ws.cell(row=row, column=4).value = record.get('one_way_distance')
                self.ws.cell(row=row, column=5).value = record.get('round_trip_distance')
                self.ws.cell(row=row, column=6).value = record.get('estimated_time')
                self.ws.cell(row=row, column=7).value = record.get('route_description')
            
            # 調整欄寬
            self.ws.column_dimensions['A'].width = 15
            self.ws.column_dimensions['B'].width = 30
            self.ws.column_dimensions['C'].width = 30
            self.ws.column_dimensions['D'].width = 12
            self.ws.column_dimensions['E'].width = 12
            self.ws.column_dimensions['F'].width = 15
            self.ws.column_dimensions['G'].width = 40
            
            return self.wb
            
        except Exception as e:
            logger.error(f"產生 Excel 報表錯誤: {str(e)}")
            raise
    
    def save(self, filepath):
        """儲存檔案"""
        self.wb.save(filepath)

class PDFReportGenerator:
    """PDF 報表產生器"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#28A745'),
            spaceAfter=30,
            alignment=1
        ))
    
    def generate_mileage_report(self, records, report_type='detail', output_path='report.pdf', include_map=False, include_route=False, include_distance=True):
        """產生里程 PDF 報表"""
        try:
            doc = SimpleDocTemplate(
                output_path,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            story = []
            
            # 標題
            title = Paragraph('地點里程報表', self.styles['CustomTitle'])
            story.append(title)
            story.append(Spacer(1, 0.2*inch))
            
            # 產生日期
            date_str = datetime.now().strftime('%Y年%m月%d日')
            date_para = Paragraph(f'產生日期：{date_str}', self.styles['Normal'])
            story.append(date_para)
            story.append(Spacer(1, 0.2*inch))
            
            # 資料表格
            data = [['日期', '起點', '終點', '單程(km)', '往返(km)', '預估時間']]
            
            for record in records:
                data.append([
                    record.get('travel_date', ''),
                    record.get('start_location', ''),
                    record.get('end_location', ''),
                    str(record.get('one_way_distance', 0)),
                    str(record.get('round_trip_distance', 0)),
                    record.get('estimated_time', '')
                ])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28A745')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            
            # 建立 PDF
            doc.build(story)
            
            logger.info(f"PDF 報表產生成功: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"產生 PDF 報表錯誤: {str(e)}")
            raise









